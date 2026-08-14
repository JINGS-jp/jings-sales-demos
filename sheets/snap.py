#!/usr/bin/env python3
"""生成した帳票を「実物の写真」PNGにして、デモ埋め込み用のdata URIを作る

処理:
  xlsx → 対象シートのみの一時xlsx（横向き・幅1ページfit・余白最小）
       → LibreOffice headless で PDF → PyMuPDF で1ページ目をPNG → PIL で白余白トリム
  すべて実際のxlsxをLibreOfficeでレンダリングしたもの（描画した絵は使わない）

出力:
  sheets/png/*.png          目視確認用
  sheets/sheets.json        {key: "data:image/png;base64,..."} ビルドが読み込む

使い方:
    python3 sheets/snap.py
"""
import base64
import io
import json
import subprocess
import sys
import tempfile
from pathlib import Path

import fitz
import openpyxl
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
SHEETS = ROOT / "sheets"
PNG = SHEETS / "png"
SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
WIDTH = 1250


def xlsx_to_png(xlsx: Path, sheet: str, area: str, out: Path):
    wb = openpyxl.load_workbook(xlsx)
    if sheet not in wb.sheetnames:
        sys.exit(f"シートが見つかりません: {sheet} / {wb.sheetnames}")
    for n in list(wb.sheetnames):
        if n != sheet:
            del wb[n]
    ws = wb[sheet]
    ws.print_area = area
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.15
    ws.page_margins.top = ws.page_margins.bottom = 0.15
    ws.page_margins.header = ws.page_margins.footer = 0
    # 「Excelで開いた画面」に見えるよう、行番号・列記号とセル罫線を出す
    ws.print_options.headings = True
    ws.print_options.gridLines = True

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / "s.xlsx"
        wb.save(tmp)
        # soffice は既定プロファイルを排他ロックするため、呼び出しごとに専用プロファイルを使う。
        # まれに失敗するので数回試す。
        pdf = Path(td) / "s.pdf"
        for attempt in range(3):
            subprocess.run([SOFFICE, f"-env:UserInstallation=file://{td}/lo{attempt}",
                            "--headless", "--norestore", "--convert-to", "pdf",
                            "--outdir", td, str(tmp)],
                           capture_output=True, timeout=240)
            if pdf.exists():
                break
        if not pdf.exists():
            sys.exit(f"PDF化に失敗: {xlsx.name}")
        doc = fitz.open(pdf)
        page = doc[0]
        zoom = WIDTH / page.rect.width
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        doc.close()

    # 白余白のトリム
    bg = Image.new("RGB", img.size, (255, 255, 255))
    from PIL import ImageChops
    diff = ImageChops.difference(img, bg).convert("L")
    box = diff.point(lambda v: 255 if v > 12 else 0).getbbox()
    if box:
        pad = 10
        img = img.crop((max(0, box[0] - pad), max(0, box[1] - pad),
                        min(img.width, box[2] + pad), min(img.height, box[3] + pad)))
    img.convert("P", palette=Image.ADAPTIVE, colors=128).save(out, optimize=True)
    return out


def main():
    PNG.mkdir(parents=True, exist_ok=True)
    man = json.loads((SHEETS / "out" / "manifest.json").read_text(encoding="utf-8"))
    js = 'const fs=require("fs");const s=fs.readFileSync(process.argv[1],"utf8");' \
         'console.log(JSON.stringify(eval(s+"; DATA")));'
    d = json.loads(subprocess.run(["node", "-e", js, str(ROOT / "src/common/data.js")],
                                  capture_output=True, text=True, check=True).stdout)

    imgs = {}
    for m in man:
        out = PNG / (m["key"] + ".png")
        xlsx_to_png(ROOT / m["file"], m["sheet"], m["area"], out)
        imgs[m["key"]] = out
        print(f"  {out.name}  {out.stat().st_size // 1024} KB")

    data = {k: "data:image/png;base64," + base64.b64encode(p.read_bytes()).decode()
            for k, p in imgs.items()}
    (SHEETS / "sheets.json").write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    total = sum(len(v) for v in data.values()) // 1024
    print(f"帳票画像 {len(data)} 件 / data URI 合計 {total} KB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
