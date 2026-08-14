#!/usr/bin/env python3
"""JINGSデモ精機の帳票を data.js の内容から生成する

デモの「出典」ボタンで実物として見せるための帳票を作る。画面のデータと帳票の中身は
どちらも src/common/data.js 由来なので、構造的にズレない。

様式は業界の標準的な構成（AIAG-VDAの列構成、苦情報告書の不良現象8分類、
Cover/改訂履歴のシート構成）に倣う。実顧客の帳票は一切参照・流用しない。

使い方:
    python3 sheets/gen_sheets.py          # sheets/out/ に xlsx を生成
"""
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "sheets" / "out"

# --- 見た目（帳票らしさ） ---
THIN = Side(style="thin", color="808080")
MED = Side(style="medium", color="404040")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="DCE6F1")
TITLE_FILL = PatternFill("solid", fgColor="F2F2F2")
MARK_FILL = PatternFill("solid", fgColor="FFF2CC")
JP = "Yu Gothic"


def load_data():
    """data.js を node 経由で読む（JSとPythonで同じ1つの正本を使う）"""
    js = 'const fs=require("fs");const s=fs.readFileSync(process.argv[1],"utf8");' \
         'console.log(JSON.stringify(eval(s+"; DATA")));'
    r = subprocess.run(["node", "-e", js, str(ROOT / "src/common/data.js")],
                       capture_output=True, text=True, check=True)
    return json.loads(r.stdout)


def style_header(ws, row, ncol, height=32):
    ws.row_dimensions[row].height = height
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = Font(name=JP, size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=MED, bottom=MED)


def put(ws, row, col, value, *, bold=False, size=9, wrap=True, fill=None,
        halign="left", valign="top", box=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=JP, size=size, bold=bold)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if box:
        cell.border = BOX
    if fill:
        cell.fill = fill
    return cell


def widths(ws, spec):
    for i, w in enumerate(spec, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_block(ws, title, meta, ncol):
    """帳票上部の表題欄。meta は (ラベル, 値) の並び。"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = put(ws, 1, 1, title, bold=True, size=14, halign="center", valign="center",
            fill=TITLE_FILL)
    t.border = Border(left=MED, right=MED, top=MED, bottom=MED)
    ws.row_dimensions[1].height = 26
    r = 2
    for i in range(0, len(meta), 2):
        pair = meta[i:i + 2]
        col = 1
        for label, value in pair:
            put(ws, r, col, label, bold=True, size=9, fill=HEAD_FILL, halign="right")
            span = (ncol // 2) - 1 if len(pair) == 2 else ncol - 1
            ws.merge_cells(start_row=r, start_column=col + 1,
                           end_row=r, end_column=col + span)
            put(ws, r, col + 1, value, size=9, halign="left")
            for c in range(col + 1, col + span + 1):
                ws.cell(row=r, column=c).border = BOX
            col += span + 1
        ws.row_dimensions[r].height = 18
        r += 1
    return r + 1


# ============================================================ ① 工程FMEA
def gen_pfmea(d):
    wb = Workbook()
    ws = wb.active
    ws.title = "1.Cover"
    widths(ws, [22, 34, 20, 30])
    r = title_block(ws, "工程FMEA　表紙", [
        ("帳票管理No.", "PFM-ACT220-001"), ("改訂No.", "Ver.09"),
        ("初版帳票作成日", "2022.09.14"), ("帳票改定日", "2026.07.09"),
        ("会社名", d["CO"]["name"]), ("工場名", "第一工場・第二工場"),
        ("部品名称", "電動アクチュエータ 高トルク品"), ("部品番号", "ACT-220"),
        ("対象ライン", "第一工場 組立ライン（工程10〜19）"), ("作成部署", "生産技術部"),
        ("作成者", "佐藤"), ("承認者", "大野"),
        ("評価基準", d["SOD_CRITERIA"]["doc"]), ("備考", "デモ環境・架空データ"),
    ], 4)
    put(ws, r, 1, "注意", bold=True, fill=MARK_FILL)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    put(ws, r, 2, "本帳票はJINGSのデモ用に生成した架空データです。実在の企業・製品とは関係ありません。")
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BOX

    # 改訂履歴
    hs = wb.create_sheet("2.改訂履歴")
    widths(hs, [10, 14, 52, 12, 12])
    r = title_block(hs, "改訂履歴", [("帳票管理No.", "PFM-ACT220-001"), ("現行版", "Ver.09")], 5)
    cols = ["改訂No.", "改定日", "改訂理由", "起案", "承認"]
    for i, c in enumerate(cols, 1):
        put(hs, r, i, c, bold=True, fill=HEAD_FILL, halign="center")
    style_header(hs, r, 5, height=20)
    hist = [
        ("Ver.09", "2026.07.09", "工程17 ねじ座面の打痕（QT-2026-0121）に伴う検出手段の見直し", "佐藤", "大野"),
        ("Ver.08", "2026.01.15", "工程18 作動音判定のばらつき（QT-2026-0012）を反映", "田中", "大野"),
        ("Ver.07", "2025.06.05", "基板固定点の変更（ECN-2025-014）を反映", "伊藤", "大野"),
        ("Ver.06", "2025.03.12", "配線ガイド形状の追加（ECN-2025-007）を反映", "鈴木", "大野"),
    ]
    for i, row in enumerate(hist):
        for j, v in enumerate(row, 1):
            put(hs, r + 1 + i, j, v, halign="center" if j != 3 else "left")

    # 本表（AIAG-VDAの列構成）
    fs = wb.create_sheet("工程FMEA")
    cols = ["工程番号", "工程名", "プロセス機能", "要求事項", "潜在的故障モード",
            "潜在的故障影響", "厳しさ\nS", "クラス", "潜在的故障原因／メカニズム",
            "発生頻度\nO", "現行プロセス管理\n予防", "現行プロセス管理\n検出",
            "検出可能性\nD", "RPN", "推奨処置", "責任者", "対策実施日"]
    widths(fs, [8, 16, 24, 26, 24, 24, 5, 7, 26, 5, 24, 22, 5, 6, 22, 8, 11])
    r = title_block(fs, "工程FMEA（様式1）", [
        ("帳票管理No./改訂No.", "PFM-ACT220-001　/　Ver.09"), ("部品番号", "ACT-220"),
        ("対象ライン", "第一工場 組立ライン"), ("帳票改定日", "2026.07.09"),
    ], 17)
    for i, c in enumerate(cols, 1):
        put(fs, r, i, c, bold=True, fill=HEAD_FILL, halign="center")
    style_header(fs, r, 17, height=42)
    fs.freeze_panes = fs.cell(row=r + 1, column=1)

    procs = {p["no"]: p for p in d["PROCESSES"]}
    reqs = d["PROC_REQS"]
    rr = r + 1
    for row in d["PFMEA"]:
        p = procs.get(row["proc"], {})
        # 要求事項は同工程の該当しそうなものを1つ充てる（帳票としての体裁）
        cand = reqs.get(row["proc"], [])
        req = cand[0]["req"] + "（" + cand[0]["spec"] + "）" if cand else ""
        rpn = row["s"] * row["o"] * row["d"]
        vals = [row["proc"], p.get("name", ""), p.get("func", ""), req, row["mode"],
                row["eff"], row["s"], p.get("mark", ""), row["cause"], row["o"],
                row["prev"], row["det"], row["d"], rpn,
                "現在の工程管理に同じ" if rpn < 100 else row["prev"],
                p.get("name", "")[:2] and "生技", "" if rpn < 100 else "2026.09.30"]
        for j, v in enumerate(vals, 1):
            put(fs, rr, j, v, halign="center" if j in (1, 7, 8, 10, 13, 14, 17) else "left",
                fill=MARK_FILL if (j == 7 and row["s"] >= 8) else None)
        fs.row_dimensions[rr].height = 34
        rr += 1

    path = OUT / "工程FMEA_ACT-220_Ver.09.xlsx"
    wb.save(path)
    return path, "工程FMEA", f"A1:Q{min(rr - 1, 22)}"


# ============================================================ ② 苦情報告書
def gen_complaint(d, tr_id="QT-2023-0187"):
    tr = next(t for t in d["TROUBLES"] if t["id"] == tr_id)
    procs = {p["no"]: p for p in d["PROCESSES"]}
    p = procs.get(tr["proc"], {})
    KINDS = ["1.異品混入", "2.寸法不良", "3.加工不良", "4.未加工不良",
             "5.外観不良", "6.品違い", "7.荷姿不良", "8.その他"]
    # 現象の記載から不良現象の区分を当てる
    sym = tr["sym"]
    pick = "8.その他"
    for key, k in [("寸法", "2.寸法不良"), ("径", "2.寸法不良"), ("トルク", "3.加工不良"),
                   ("バリ", "3.加工不良"), ("打痕", "5.外観不良"), ("錆", "5.外観不良"),
                   ("嵌合", "4.未加工不良"), ("異音", "8.その他"), ("漏", "8.その他")]:
        if key in sym:
            pick = k
            break

    wb = Workbook()
    ws = wb.active
    ws.title = "苦情報告書A"
    widths(ws, [13, 20, 13, 20, 13, 20, 13, 20])
    r = title_block(ws, "苦情報告書（品質苦情処理規定 QR-2201）", [
        ("苦情No.", tr["id"]), ("発行日", tr["date"]),
        ("発生日", tr["date"]), ("ランク", "A" if tr["leak"] else "B"),
        ("顧客名", "顧客A社" if tr["leak"] else "（社内検出）"),
        ("工場", "第一工場" if p.get("plant") == "P1" else "第二工場"),
        ("品番", tr["prod"]), ("発生工程", f'工程{tr["proc"]} {p.get("name","")}'),
        ("発生責任部署", "生産技術部"), ("担当者", tr["owner"]),
        ("対象部位", tr["part"]), ("社外流出", "あり" if tr["leak"] else "なし"),
    ], 8)

    def block(label, text, height=44, fill=None):
        nonlocal r
        put(ws, r, 1, label, bold=True, fill=fill or HEAD_FILL, halign="center",
            valign="center")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        put(ws, r, 2, text)
        for c in range(2, 9):
            ws.cell(row=r, column=c).border = BOX
        ws.row_dimensions[r].height = height
        r += 1

    block("発見の経緯", "顧客ラインでの組付時に発見" if tr["leak"] else "社内の検査工程で発見", 20)
    block("不良現象", sym)

    # 不良現象の区分（8分類のチェック欄）
    put(ws, r, 1, "不良現象の区分", bold=True, fill=HEAD_FILL, halign="center", valign="center")
    # 4件ずつ2行。各件を列2つ分に割り当てて幅いっぱいに収める（空セルを残さない）
    SPANS = [(2, 3), (4, 5), (6, 7), (8, 8)]
    for i, k in enumerate(KINDS):
        rowx = r + (i // 4)
        c0, c1 = SPANS[i % 4]
        if c1 > c0:
            ws.merge_cells(start_row=rowx, start_column=c0, end_row=rowx, end_column=c1)
        mark = "■ " if k == pick else "□ "
        put(ws, rowx, c0, mark + k, halign="center", valign="center",
            bold=(k == pick), fill=MARK_FILL if k == pick else None)
        for c in range(c0, c1 + 1):
            ws.cell(row=rowx, column=c).border = BOX
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    for i in range(2):
        ws.row_dimensions[r + i].height = 20
    r += 2

    block("初期情報・暫定処置", tr["tmp"])
    block("原因（発生原因）", tr["cause"])
    block("原因（流出原因）",
          "現行の検出手段では本件を検出できなかった。検出をすり抜けた条件は調査中。" if tr["leak"]
          else "社内検査で検出済み。流出なし。", 32)
    block("恒久対策", tr["perm"])

    put(ws, r, 1, "評価", bold=True, fill=HEAD_FILL, halign="center")
    for i, (lab, v) in enumerate([("影響度S", tr["s"]), ("発生度O", tr["o"]), ("検出度D", tr["d"])]):
        put(ws, r, 2 + i * 2, lab, bold=True, halign="right")
        put(ws, r, 3 + i * 2, v, halign="center",
            fill=MARK_FILL if lab == "影響度S" and tr["s"] >= 8 else None)
    put(ws, r, 8, f'RPN {tr["s"] * tr["o"] * tr["d"]}', halign="center", bold=True)
    ws.row_dimensions[r].height = 20
    r += 1
    block("対応状態", tr["status"], 20)
    block("備考", "本帳票はJINGSのデモ用に生成した架空データです。実在の企業・製品とは関係ありません。", 20)

    # 品質会議資料（2枚目）
    qs = wb.create_sheet("品質会議資料")
    widths(qs, [14, 30, 14, 30])
    r2 = title_block(qs, "品質会議資料（苦情報告書 添付）", [
        ("苦情No.", tr["id"]), ("報告日", tr["date"])], 4)
    for lab, v in [("発生事象", sym), ("原因", tr["cause"]), ("恒久対策", tr["perm"]),
                   ("水平展開の範囲", "同一工程の他機種、および同一原因の他工程を確認する"),
                   ("再発防止の確認方法", "対策後3か月の発生件数を監視する")]:
        put(qs, r2, 1, lab, bold=True, fill=HEAD_FILL, halign="center", valign="center")
        qs.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=4)
        put(qs, r2, 2, v)
        for c in range(2, 5):
            qs.cell(row=r2, column=c).border = BOX
        qs.row_dimensions[r2].height = 40
        r2 += 1

    path = OUT / f"苦情報告書_{tr_id}_{tr['prod']}.xlsx"
    wb.save(path)
    return path, "苦情報告書A", f"A1:H{r - 1}"


# ============================================================ ③ 変更発議書
def gen_ecr(d):
    wb = Workbook()
    ws = wb.active
    ws.title = "変更発議書"
    widths(ws, [6, 30, 26, 26, 12, 14])
    r = title_block(ws, "設計変更発議書", [
        ("発議No.", "ECR-2026-014"), ("発議日", "2026.07.01"),
        ("対象品番", "ACT-230（電動アクチュエータ 小型軽量品）"), ("現行品番", "ACT-220"),
        ("発議部署", "技術部"), ("発議者", "森"),
        ("変更理由", "小型軽量化と応答速度の向上、および耐熱性の確保"),
        ("量産予定", "2027.01"),
    ], 6)
    cols = ["No.", "変更点", "変更理由", "変更の目的・ねらい", "影響範囲", "確認担当"]
    for i, c in enumerate(cols, 1):
        put(ws, r, i, c, bold=True, fill=HEAD_FILL, halign="center")
    style_header(ws, r, 6, height=24)

    seen, rr = [], r + 1
    for row in d["DRBFM"]:
        if row["cp"] in seen:
            continue
        seen.append(row["cp"])
        scope = "工程FMEA・図面・作業要領書" if "締結" in row["cp"] or "ハウジング" in row["cp"] \
            else "工程FMEA・図面"
        for j, v in enumerate([len(seen), row["cp"], row["why"],
                               "現行機種からの変更により品質への影響を確認する",
                               scope, "森"], 1):
            put(ws, rr, j, v, halign="center" if j in (1, 6) else "left")
        ws.row_dimensions[rr].height = 34
        rr += 1

    put(ws, rr + 1, 1, "承認欄", bold=True, fill=HEAD_FILL, halign="center")
    for i, (lab, who) in enumerate([("起案", "森"), ("審査", "中村"), ("承認", "大野")]):
        put(ws, rr + 1, 2 + i, f"{lab}：{who}", halign="center")
    put(ws, rr + 3, 1, "備考", bold=True, fill=HEAD_FILL, halign="center")
    ws.merge_cells(start_row=rr + 3, start_column=2, end_row=rr + 3, end_column=6)
    put(ws, rr + 3, 2, "本帳票はJINGSのデモ用に生成した架空データです。実在の企業・製品とは関係ありません。")
    for c in range(2, 7):
        ws.cell(row=rr + 3, column=c).border = BOX

    path = OUT / "変更発議書_ACT-230_20260701.xlsx"
    wb.save(path)
    return path, "変更発議書", f"A1:F{rr + 3}"


# ============================================================ ⑤ 工程フロー・QC工程表
def gen_flow(d):
    wb = Workbook()
    ws = wb.active
    ws.title = "工程フロー図"
    widths(ws, [8, 22, 8, 22, 10, 26, 14])
    r = title_block(ws, "工程フロー図", [
        ("帳票管理No.", "PF-ACT220-001"), ("改訂No.", "Ver.06"),
        ("対象品番", "ACT-220"), ("対象ライン", "第一工場 組立ライン"),
    ], 7)
    cols = ["工程番号", "工程名", "工程記号", "設備・治具", "特性記号", "工程の機能", "工場"]
    for i, c in enumerate(cols, 1):
        put(ws, r, i, c, bold=True, fill=HEAD_FILL, halign="center")
    style_header(ws, r, 7, height=22)
    EQ = {"10": "射出成形機 IM-150", "11": "圧入プレス PP-20", "12": "組付治具 JG-12",
          "13": "位置決め治具 JG-13", "14": "リフロー炉 RF-30", "15": "配線ガイド治具 JG-15",
          "16": "定量吐出装置 GD-16", "17": "電動ドライバ ED-17", "18": "機能検査機 FT-18",
          "19": "気密検査機 LT-19"}
    for i, p in enumerate(d["PROCESSES"]):
        vals = [p["no"], p["name"], "○" if i else "▷", EQ.get(p["no"], ""), p["mark"],
                p["func"], "第一工場" if p["plant"] == "P1" else "第二工場"]
        for j, v in enumerate(vals, 1):
            put(ws, r + 1 + i, j, v, halign="center" if j in (1, 3, 5, 7) else "left",
                fill=MARK_FILL if j == 5 and p["mark"] else None)
        ws.row_dimensions[r + 1 + i].height = 26

    qs = wb.create_sheet("QC工程表")
    widths(qs, [8, 20, 26, 22, 18, 14, 16, 14])
    r2 = title_block(qs, "QC工程表", [
        ("帳票管理No.", "QC-ACT220-001"), ("改訂No.", "Ver.06"),
        ("対象品番", "ACT-220"), ("作成部署", "生産技術部"),
    ], 8)
    cols2 = ["工程番号", "工程名", "管理項目", "規格値・条件", "管理方法",
             "測定頻度", "記録", "特性記号"]
    for i, c in enumerate(cols2, 1):
        put(qs, r2, i, c, bold=True, fill=HEAD_FILL, halign="center")
    style_header(qs, r2, 8, height=22)
    rr = r2 + 1
    procs = {p["no"]: p for p in d["PROCESSES"]}
    for no, items in d["PROC_REQS"].items():
        p = procs.get(no, {})
        for it in items:
            auto = "自動測定・全数" if "記録" in it.get("spec", "") or it["conf"] == "doc" else "目視・全数"
            vals = [no, p.get("name", ""), it["req"], it["spec"], auto,
                    "全数" if "全数" in auto else "1回/ロット",
                    "測定値をライン側で記録", p.get("mark", "")]
            for j, v in enumerate(vals, 1):
                put(qs, rr, j, v, halign="center" if j in (1, 6, 8) else "left",
                    fill=MARK_FILL if j == 8 and p.get("mark") else None)
            qs.row_dimensions[rr].height = 28
            rr += 1

    path = OUT / "工程フロー_QC工程表_ACT-220.xlsx"
    wb.save(path)
    return path, "工程フロー図", f"A1:G{r + len(d['PROCESSES'])}"


# ============================================================ ⑥ 設計審査記録
def gen_dr(d, gate="DR3"):
    g = next(x for x in d["DR_GATES"] if x["id"] == gate)
    wb = Workbook()
    ws = wb.active
    ws.title = "設計審査記録"
    widths(ws, [10, 12, 48, 12, 20, 14, 12])
    r = title_block(ws, f"設計審査記録（{g['id']} {g['name']}）", [
        ("帳票管理No.", f"DR-ACT230-{gate}"), ("対象品番", "ACT-230"),
        ("実施予定日", g["date"]), ("主査", g["chair"]),
        ("審査の目的", g["note"]), ("確認項目数", f'{len(d["DR_CHECKLIST"][gate])} 件'),
    ], 7)
    cols = ["項目番号", "区分", "確認項目", "判定", "指摘・コメント", "担当", "期限"]
    for i, c in enumerate(cols, 1):
        put(ws, r, i, c, bold=True, fill=HEAD_FILL, halign="center")
    style_header(ws, r, 7, height=22)

    fnd = {f["item"]: f for f in d["DR_FINDINGS"] if f["gate"] == gate}
    rr = r + 1
    for it in d["DR_CHECKLIST"][gate]:
        f = fnd.get(it["item"])
        vals = [it["id"], it["cat"], it["item"],
                "指摘" if f else "未実施",
                f["item"] if f else "", f["by"] if f else "", f["due"] if f else ""]
        for j, v in enumerate(vals, 1):
            put(ws, rr, j, v, halign="center" if j in (1, 2, 4, 7) else "left",
                fill=MARK_FILL if j == 4 and f else None)
        ws.row_dimensions[rr].height = 30
        rr += 1

    put(ws, rr + 1, 1, "備考", bold=True, fill=HEAD_FILL, halign="center")
    ws.merge_cells(start_row=rr + 1, start_column=2, end_row=rr + 1, end_column=7)
    put(ws, rr + 1, 2, "本帳票はJINGSのデモ用に生成した架空データです。実在の企業・製品とは関係ありません。")
    for c in range(2, 8):
        ws.cell(row=rr + 1, column=c).border = BOX

    path = OUT / f"設計審査記録_{gate}_ACT-230.xlsx"
    wb.save(path)
    return path, "設計審査記録", f"A1:G{rr - 1}"


# ============================================================ ④ 図面属性表
def gen_dwg(d, no="ACT-230-300"):
    """図面から読み取った寸法・公差・注記・部品表。検図はこの内容に対して行う。
    図面そのものの画像は作らない（実物のCAD/PDFがないため、絵を描いて実物のように
    見せることはしない）。"""
    dw = next(x for x in d["DRAWINGS"] if x["no"] == no)
    hits = d["DWG_FINDINGS"].get(no, [])
    rules = {r["id"]: r for r in d["DWG_RULES"]}

    wb = Workbook()
    ws = wb.active
    ws.title = "図面属性表"
    widths(ws, [8, 26, 22, 18, 16, 30])
    r = title_block(ws, f"図面属性表（{no}）", [
        ("図面番号", dw["no"]), ("版", dw["rev"]),
        ("名称", dw["name"]), ("対象機種", dw["prod"]),
        ("発行日", dw["date"]), ("尺度／投影法", "1:1　／　第三角法"),
        ("設計", "森"), ("審査／承認", "（未記入）"),
        ("材質", "PPS（GF40%）"), ("特性記号", "（指定なし）"),
        ("改訂理由", "（記載なし）"), ("備考", "デモ環境・架空データ"),
    ], 6)

    def section(label, cols, rows):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        put(ws, r, 1, label, bold=True, size=11, fill=TITLE_FILL, halign="left")
        r += 1
        for i, c in enumerate(cols, 1):
            put(ws, r, i, c, bold=True, fill=HEAD_FILL, halign="center")
        style_header(ws, r, 6, height=20)
        r += 1
        for row in rows:
            for j, v in enumerate(row, 1):
                put(ws, r, j, v, halign="center" if j == 1 else "left")
            ws.row_dimensions[r].height = 24
            r += 1
        r += 1

    section("寸法・公差", ["No.", "対象部位", "寸法", "公差", "等級", "備考"], [
        ("1", "外形（幅）", "86.0", "±0.2", "—", ""),
        ("2", "外形（高さ）", "54.0", "±0.2", "—", ""),
        ("3", "軸受穴", "φ32.0", "+0.03 / 0", "H7相当", "モータ軸との嵌合部"),
        ("4", "ガスケット溝深さ", "1.5", "±0.05", "—", "相手部品側の公差は未指定"),
        ("5", "板厚", "3.0", "（指示なし）", "—", "公差の指定がありません"),
        ("6", "歯面幅", "—", "—", "—", "本図の対象外"),
    ])
    section("注記", ["No.", "内容", "", "", "", ""], [
        ("1", "指示なき寸法公差は JIS B 0405-m による。", "", "", "", ""),
        ("2", "ガスケット溝の面粗さは Ra 3.2 以下とする。", "", "", "", ""),
        ("3", "ケース締結は M3 × 6本とする（部品表 項目8）。", "", "", "", ""),
        ("4", "材質 PPS（GF40%）。", "", "", "", ""),
        ("5", "締付順序の指示は本図には記載していない。", "", "", "", ""),
    ])
    section("部品表", ["項目", "名称", "材質", "員数", "図中の指示", "備考"], [
        ("1", "ケース", "PPS（GF40%）", "1", "—", ""),
        ("8", "締結ねじ M3×8", "SUS304", "6", "4箇所のみ記入", "員数と図中指示が不一致"),
        ("9", "ガスケット", "シリコーン", "1", "断面A-A", ""),
    ])
    section("改訂履歴", ["改訂No.", "改定日", "改訂理由", "起案", "審査", "承認"], [
        ("A", dw["date"], "（記載なし）", "森", "（未記入）", "（未記入）"),
    ])

    ws2 = wb.create_sheet("検図チェック記録")
    widths(ws2, [10, 10, 22, 34, 30, 12])
    r2 = title_block(ws2, "検図チェック記録", [
        ("図面番号", dw["no"]), ("版", dw["rev"]),
        ("検図者", "（未実施）"), ("検図日", "（未実施）"),
    ], 6)
    for i, c in enumerate(["ルール", "重要度", "該当箇所", "確認内容", "理由", "判定"], 1):
        put(ws2, r2, i, c, bold=True, fill=HEAD_FILL, halign="center")
    style_header(ws2, r2, 6, height=20)
    for i, h in enumerate(hits):
        vals = [h["rule"], h["sev"], h["where"], h["found"], h["why"], "未判定"]
        for j, v in enumerate(vals, 1):
            put(ws2, r2 + 1 + i, j, v,
                halign="center" if j in (1, 2, 6) else "left",
                fill=MARK_FILL if j == 2 and h["sev"] == "重" else None)
        ws2.row_dimensions[r2 + 1 + i].height = 32

    path = OUT / f"図面属性表_{no}_Rev{dw['rev']}.xlsx"
    wb.save(path)
    return path, "図面属性表", f"A1:F{r - 2}"


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    d = load_data()
    made = []
    made.append(("pfmea",) + gen_pfmea(d))
    made.append(("complaint",) + gen_complaint(d))
    made.append(("ecr",) + gen_ecr(d))
    made.append(("flow",) + gen_flow(d))
    made.append(("dr",) + gen_dr(d))
    made.append(("drawing",) + gen_dwg(d))
    # 苦情報告書は、デモで実際に押される記録を網羅する
    # （ダッシュボードの要対応・検索例文の上位ヒット・デモ6の選択候補）
    for tid in ["QT-2025-0344", "QT-2025-0378", "QT-2026-0012", "QT-2026-0038", "QT-2026-0061", "QT-2026-0084", "QT-2026-0103", "QT-2026-0121", "QT-2024-0209", "QT-2024-0312", "QT-2023-0142", "QT-2025-0077", "QT-2023-0301"]:
        made.append((f"complaint_{tid}",) + gen_complaint(d, tid))

    manifest = OUT / "manifest.json"
    manifest.write_text(json.dumps(
        [{"key": k, "file": str(p.relative_to(ROOT)), "sheet": s, "area": a}
         for k, p, s, a in made], ensure_ascii=False, indent=2), encoding="utf-8")
    for k, p, s, a in made:
        print(f"  {p.name}  [{s}] {a}")
    print(f"帳票 {len(made)} 件 / {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
